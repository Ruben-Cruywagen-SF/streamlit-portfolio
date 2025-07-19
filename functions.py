from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
from io import BytesIO
import google.generativeai as genai
import pandas as pd
import random, os
from faker import Faker
from datetime import datetime, timedelta


def generate_dummy_data() -> pd.DataFrame:
    '''
    Generates dummy data that is compatible with the dashboard. To make it interesting, several random modifiers are used.
    Returns a pandas DataFrame.
    '''

    fake = Faker()

    num_days = 60
    max_sales_per_day = 100

    regions = []
    for region in ['North', 'South', 'East', 'West']:
        for i in range(1, random.randint(2,20)):
            regions.append(region)

    products = []
    for product in ['Widget A', 'Widget B', 'Widget C', 'Gadget X', 'Gadget Y', 'Gadget Z']:
        for i in range(1, random.randint(2,20)):
            products.append(product)

    names_unique = [fake.first_name() for _ in range(6)]
    reps = {}
    for name in names_unique:
        modifier1 =  random.random()
        modifier2 =  random.random()
        reps[name] = [min(modifier1, modifier2), max(modifier1, modifier2)] 

    names = []
    for name in names_unique:
        for i in range(1, random.randint(2,20)):
            names.append(name)

    # Generate rows
    daily_sales_modifier = random.random()
    rows = []
    start_date = datetime.today() - timedelta(days=num_days)
    for i in range(num_days):
        current_date = start_date + timedelta(days=i)
        person = random.choice(names)
        for _ in range(0, int(max_sales_per_day * daily_sales_modifier)):
            row = {
                'Date': current_date.strftime('%Y-%m-%d'),
                'Region': random.choice(regions),
                'Rep': person,
                'Product': random.choice(products),
                'Sales': round(random.uniform(0, 1000) * random.uniform(reps[person][0], reps[person][1]) , 2)
            }
            rows.append(row)
        
        if random.random() < 0.2:
            daily_sales_modifier = random.random()

    df = pd.DataFrame(rows)
    
    return df


def generate_excel_report(df:pd.DataFrame) -> BytesIO:
    '''
    Uses filtered DataFrame to populate the template Excel workbook.
    Returns as a BytesIO buffer of the workbook.
    '''

    workbook = load_workbook('sales_report_template.xlsx')
    summary_sheet = workbook['Summary'] 
    dataSheet = workbook['Filtered Data'] 

    # Summary
    summary_sheet['B21'] = datetime.strftime(datetime.now(), '%y-%m-%d %H:%M')
    summary_sheet['B22'] = ', '.join(df['Region'].unique())
    summary_sheet['B23'] = ', '.join(df['Rep'].unique())
    minDate = df['Date'].min()
    maxDate = df['Date'].max()
    summary_sheet['B24'] = f"{datetime.strftime(minDate, '%y-%m-%d')} to {datetime.strftime(maxDate, '%y-%m-%d')}"
    summary_sheet['B25'] = sum(df['Sales'])
    summary_sheet['B26'] = float(df.groupby(df['Date'])['Sales'].sum().mean())

    total_rows = len(df)
    dataSheet.insert_rows(2, amount= total_rows - 1)
    for i,r in df.iterrows():
        for col in range(len(r)):
            dataSheet[f'{get_column_letter(col+1)}{i+2}'] = r.iloc[col]
    dataSheet[f'E{total_rows+2}'] = f'=SUM(E2:E{total_rows+1})'

    # data types
    for row in dataSheet.iter_rows(min_row=2, max_row=total_rows+1, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return buffer


def generate_ai_summary(df:pd.DataFrame) -> str:
    '''
    Generates an AI summary of the provided data from the totals and averages of the columns of interest.
    Returns as the response as a string.
    '''

    ai_info = {}
    for col in ['Region', 'Rep', 'Product']:
        ai_info[col] = {}
        ai_info[col]['Total'] = df[[col,'Sales']].groupby([col]).sum().round(1).to_dict()['Sales']
        ai_info[col]['Average'] = df[[col,'Sales']].groupby([col]).mean().round(1).to_dict()['Sales']

    print(ai_info)

    minDate = df['Date'].min()
    maxDate = df['Date'].max()    

    genai.configure(api_key=os.getenv('GEMINI_KEY'))
    model = genai.GenerativeModel('gemini-2.5-flash')
    response = model.generate_content(f"You're an AI assistant summarizing sales performance. Give 1 short paragraph of insight based on this breakdown by region, rep, and product. Data period: {minDate} to {maxDate}. Respond only with the paragraph.\n\n{ai_info}")
    
    return response.text
